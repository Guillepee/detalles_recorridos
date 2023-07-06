# -*- coding: utf-8 -*-
"""
Created on Thu Jun 22 10:46:47 2023
@author: Guillermo.Palmieri23@gmail.com

Este Script sirve para unificar varios detalles de corrida en uno solo, y adicionalmente realiza los siguientes calculos:

Columna N: El tiempo entre los el reporte actual y el inmediato anterior
Columna P: La progresiva del reporte anterior
Columna P: El tiempo total del viaje
Columna Q: Acumulado de tiempo sin reportar (Excluye reportes de trene detenenido, reportes Offline y reportes menores a la tolerancia)
Columna R: Acumulado de Metros sin reportar (Excluye reportes de trene detenenido, reportes Offline y reportes menores a la tolerancia)
"""

import openpyxl
import string
import os
import glob
import xlrd
import datetime
import numpy as np
import pandas as pd
from datetime import date, time, datetime, timedelta
from xlrd import open_workbook
from openpyxl import Workbook, load_workbook, cell 
from openpyxl.chart import title
from openpyxl.utils import FORMULAE
from openpyxl.styles import PatternFill
import plotly.express as px

#Aca empiezo a trabajar con todos los archivos que estan en el directorio. 

carpeta = os.listdir('.')

titulos_basedatos = ["N°","Fecha GPS","Fecha Reg.","Equipo","Velocidad","Velocidad Max.","ID Tramo","ID Referencia","Progresiva [m]","Referencia Orden","Punto",
                     "Latitud","Longitud","location","Tiempo sin reportar","Progresiva Anterior","Tiempo de Viaje","Tiempo Total sin Reporta","Metros sin Reporte"]

titulos_basedatos_calculos = ['Mat. Rodante','Viajes','Cantidad reportes tardios','Mts sin Reportar','kms sin reportar','Promedio reportes tardios x Tren',
                              'Tiempo sin reportar','Tiempo sin reportar x Tren']

class Excel:
    def __init__(self): #Se crea el arhivo base de datos con las pestañas y columnas que quiero
        
        self.archivos_excel = glob.glob('*.xlsx') #Crea una lista con todos los archicos excels en la ubicacion de este script
        self.datos_excel = []
        self.archivo = 1
        self.Iterar_excels()
        self.crear_base_unificada()

    def crear_base_unificada(self):                            
        datos_combinados = pd.concat(self.datos_excel, ignore_index=True)
        datos_combinados.to_excel('.\\Corridas Unificadas.xlsx',sheet_name="", index=False)
        
    def Iterar_excels(self):
        """
        Agrego columnas calculadas con OpenPyXl con informacion vinculada a demoras en los registros
        """
        for a in self.archivos_excel:
            self.archivo_origen = load_workbook(f'.\\{a}')
            self.hoja_origen = self.archivo_origen.active
    
            self.maxcolB_origen = self.hoja_origen.max_row # Obtengo el valor maximo de las filas en archivo origen.
            print(self.maxcolB_origen)
            
            # Establesco el rango de filas en donde quiero trabajar: De A3 hasta Axxx (maximo de filas establecido por maxcolB_origen)        
            self.rango = list(range(3,self.maxcolB_origen))
            self.tiempoviaje = timedelta(minutes=0,seconds=0)
            # Tolerancia de tiempo sin reportar 
            self.tiempo_reporte_max = timedelta(minutes=0,seconds=59)
    
            # COLUMNA N: Tiempo de reporte (C3 - C2)
            self.chequeo = 0
            self.hoja_origen['n1'] = "Tiempo Reporte"
            for i in self.rango:
                self.hoja_origen[f'n{i}'] = (self.hoja_origen.cell(row = i, column = 2).value) - (self.hoja_origen.cell(row = i -1, column = 2).value)
                self.chequeo +=1
    
            # COLUMNA O : Progresiva anterior
            self.hoja_origen['O1'] = "Prog. Anterior"
            for i in self.rango:
                self.hoja_origen[f'O{i}'] = self.hoja_origen.cell(row = i-1, column = 8).value
    
            # COLUMNA P : Tiempo de viaje total   
            self.hoja_origen['P1'] = "Tiempo de Viaje"
            for i in self.rango:
                self.hoja_origen[f'P{i}'] = self.hoja_origen.cell(row = self.maxcolB_origen-1, column = 2).value - self.hoja_origen.cell(row = 2, column = 2).value
    
            # COLUMNA Q : Acumulado de Tiempo Sin Reportar. Primero sumo en "Tiempoviaje" todos los reportes mayores a 59 segundos y mayores a 0Km/h.
            # Luego agrego en esa columna el tiempo total sin reportar.
            self.hoja_origen['Q1'] = "Tiempo sin reportar"
            for i in self.rango:
                if self.hoja_origen[f'N{i}'].value > self.tiempo_reporte_max:
                    if self.hoja_origen[f'D{i}'].value > 0:
                        self.tiempoviaje += self.hoja_origen[f'N{i}'].value
            for i in self.rango:
                self.hoja_origen[f'Q{i}'] = self.tiempoviaje
    
            # COLUMNA R: T"Metros sin Reporte"   
            self.hoja_origen['R1'] = "Metros sin Reporte"
            for i in self.rango:
                self.hoja_origen[f'R{i}'] = abs(self.hoja_origen.cell(row = i, column = 8).value - self.hoja_origen.cell(row = i, column = 15).value)
    
            # Limite de tiempo de reporte
            self.tiempo_reporte_max = timedelta(minutes=0,seconds=59)
    
            self.archivo_origen.save(f'.\\{a}_Procesado.xlsx')
            corrida = pd.read_excel(f'.\\{a}_Procesado.xlsx')
            self.datos_excel.append(corrida)
            self.mapear(corrida)

            self.archivo += 1
            print(self.archivo)
        
    def latitudes(self,latitud): #Agrega la , despues del 3er caracter.
        nueva_latitud = str(latitud).replace(".","")
        return float(nueva_latitud[0:3] + "." + nueva_latitud[3:99])      
        
    def mapear(self,corrida):
               
        self.data = corrida
        self.data["Latitud"] = self.data["Latitud"].apply(self.latitudes)
        self.data["Longitud"] = self.data["Longitud"].apply(self.latitudes)
        
        fig = px.scatter_mapbox(self.data, lat="Latitud", lon="Longitud", hover_name="Fecha GPS" , hover_data=["Progresiva [m]","Velocidad","Velocidad Max."],
                                color="Velocidad", color_continuous_scale="Reds", zoom=10)
        
        fig.update_layout(mapbox_style="open-street-map")
        fig.update_layout(title_text="Detalle de Corrida en Mapa")
        fig.update_layout(legend_title="Velocidad (km/h)")
        
        fig.write_html(f'Equipo {self.data["Equipo"][1]}.html')
        fig.show()        
        
excel = Excel()
print("Finalizado")

   